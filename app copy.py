# -*- coding: utf-8 -*-
from telegram_bot import get_telegram_bot, check_expiring_members
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from functools import wraps
from datetime import datetime, timedelta
import random
import string
import os
import qrcode
from io import BytesIO
import base64

app = Flask(__name__)

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///gym.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['TEMPLATES_AUTO_RELOAD'] = True

from models import db, Member, User, Attendance

db.init_app(app)

# Decorator untuk protect routes
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Silakan login terlebih dahulu!', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def generate_member_id():
    """Generate ID Member otomatis format: MG + 6 digit random"""
    try:
        with app.app_context():
            while True:
                new_id = 'MG' + ''.join(random.choices(string.digits, k=6))
                existing = Member.query.filter_by(member_id=new_id).first()
                if not existing:
                    return new_id
    except Exception as e:
        print(f"Error generating member ID: {e}")
        return 'MG' + ''.join(random.choices(string.digits, k=6))

def calculate_expire_date(tanggal_daftar, type_member):
    """Hitung tanggal expire berdasarkan type member"""
    if type_member == 'bulanan':
        return tanggal_daftar + timedelta(days=30)
    elif type_member == '3_bulan':
        return tanggal_daftar + timedelta(days=90)
    elif type_member == '6_bulan':
        return tanggal_daftar + timedelta(days=180)
    elif type_member == 'tahunan':
        return tanggal_daftar + timedelta(days=365)
    return tanggal_daftar + timedelta(days=30)

def generate_qr_code_base64(member_id):
    """Generate QR Code dan return base64 string"""
    try:
        qr_data = f"GYM-{member_id}"
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        img_io = BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        
        img_base64 = base64.b64encode(img_io.getvalue()).decode()
        return img_base64
    except Exception as e:
        print(f"Error generating QR code: {e}")
        return None

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['nama_lengkap'] = user.nama_lengkap
            
            user.last_login = datetime.now()
            db.session.commit()
            
            flash(f'Selamat datang, {user.nama_lengkap}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Username atau password salah!', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    nama = session.get('nama_lengkap', 'User')
    session.clear()
    flash(f'Berhasil logout. Sampai jumpa, {nama}!', 'success')
    return redirect(url_for('login'))

@app.route('/register-admin', methods=['GET', 'POST'])
def register_admin():
    admin_exists = User.query.first()
    
    if admin_exists:
        flash('Admin sudah terdaftar! Silakan login.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        nama_lengkap = request.form.get('nama_lengkap')
        email = request.form.get('email')
        
        if password != confirm_password:
            flash('Password tidak cocok!', 'error')
            return render_template('register_admin.html')
        
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Username sudah digunakan!', 'error')
            return render_template('register_admin.html')
        
        existing_email = User.query.filter_by(email=email).first()
        if existing_email:
            flash('Email sudah digunakan!', 'error')
            return render_template('register_admin.html')
        
        new_admin = User(
            username=username,
            nama_lengkap=nama_lengkap,
            email=email,
            role='admin'
        )
        new_admin.set_password(password)
        
        db.session.add(new_admin)
        db.session.commit()
        
        flash('Admin berhasil didaftarkan! Silakan login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register_admin.html')

@app.route('/')
@login_required
def index():
    try:
        today = datetime.now()
        member_aktif = Member.query.filter(Member.tanggal_expire >= today).count()
        member_expired = Member.query.filter(Member.tanggal_expire < today).count()
        total_pendapatan = db.session.query(db.func.sum(Member.total)).scalar() or 0
        
        tiga_hari_kedepan = today + timedelta(days=3)
        member_akan_expired = Member.query.filter(
            Member.tanggal_expire >= today,
            Member.tanggal_expire <= tiga_hari_kedepan
        ).order_by(Member.tanggal_expire.asc()).all()
        
        list_member_aktif = Member.query.filter(
            Member.tanggal_expire >= today
        ).order_by(Member.tanggal_expire.asc()).all()
        
        return render_template('base.html', 
                             member_aktif=member_aktif,
                             member_expired=member_expired,
                             total_pendapatan=total_pendapatan,
                             member_akan_expired=member_akan_expired,
                             list_member_aktif=list_member_aktif,
                             today=today)
    except Exception as e:
        print(f"Error in index: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    try:
        if request.method == 'POST':
            type_member = request.form.get('type_member')
            biaya_bulanan = float(request.form.get('biaya_bulanan', 0))
            biaya_pendaftaran = float(request.form.get('biaya_pendaftaran', 0))
            nama = request.form.get('nama', '')
            alamat = request.form.get('alamat', '')
            jenis_kelamin = request.form.get('jenis_kelamin', '')
            pekerjaan = request.form.get('pekerjaan', '')
            no_handphone = request.form.get('no_handphone', '')
            
            if not all([type_member, nama, alamat, jenis_kelamin, pekerjaan, no_handphone]):
                flash('Semua field harus diisi!', 'error')
                return render_template('register.html')
            
            member_id = generate_member_id()
            tanggal_daftar = datetime.now()
            tanggal_expire = calculate_expire_date(tanggal_daftar, type_member)
            
            if type_member == 'bulanan':
                total = biaya_bulanan + biaya_pendaftaran
            elif type_member == '3_bulan':
                total = (biaya_bulanan * 3) + biaya_pendaftaran
            elif type_member == '6_bulan':
                total = (biaya_bulanan * 6) + biaya_pendaftaran
            elif type_member == 'tahunan':
                total = (biaya_bulanan * 12) + biaya_pendaftaran
            else:
                total = biaya_bulanan + biaya_pendaftaran
            
            new_member = Member(
                member_id=member_id,
                tanggal_daftar=tanggal_daftar,
                tanggal_expire=tanggal_expire,
                type_member=type_member,
                biaya_bulanan=biaya_bulanan,
                biaya_pendaftaran=biaya_pendaftaran,
                total=total,
                nama=nama,
                alamat=alamat,
                jenis_kelamin=jenis_kelamin,
                pekerjaan=pekerjaan,
                no_handphone=no_handphone
            )
            
            db.session.add(new_member)
            db.session.commit()
            
# 🔔 TAMBAHKAN INI - Send Telegram notification
            try:
                bot = get_telegram_bot()
                if bot:
                    bot.send_new_member_alert(new_member)
            except Exception as e:
                print(f"Failed to send telegram notification: {e}")
            
            flash(f'Member berhasil terdaftar dengan ID: {member_id}', 'success')
            return redirect(url_for('success', member_id=member_id, auto_qr='1'))
        
        return render_template('register.html')
    
    except Exception as e:
        print(f"Error in register: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/edit/<member_id>', methods=['GET', 'POST'])
@login_required
def edit_member(member_id):
    try:
        member = Member.query.filter_by(member_id=member_id).first()
        
        if not member:
            flash('Member tidak ditemukan!', 'error')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            member.nama = request.form.get('nama', '')
            member.alamat = request.form.get('alamat', '')
            member.jenis_kelamin = request.form.get('jenis_kelamin', '')
            member.pekerjaan = request.form.get('pekerjaan', '')
            member.no_handphone = request.form.get('no_handphone', '')
            member.type_member = request.form.get('type_member')
            member.biaya_bulanan = float(request.form.get('biaya_bulanan', 0))
            member.biaya_pendaftaran = float(request.form.get('biaya_pendaftaran', 0))
            
            tanggal_daftar = member.tanggal_daftar
            member.tanggal_expire = calculate_expire_date(tanggal_daftar, member.type_member)
            
            if member.type_member == 'bulanan':
                member.total = member.biaya_bulanan + member.biaya_pendaftaran
            elif member.type_member == '3_bulan':
                member.total = (member.biaya_bulanan * 3) + member.biaya_pendaftaran
            elif member.type_member == '6_bulan':
                member.total = (member.biaya_bulanan * 6) + member.biaya_pendaftaran
            elif member.type_member == 'tahunan':
                member.total = (member.biaya_bulanan * 12) + member.biaya_pendaftaran
            
            db.session.commit()
            
            flash(f'Data member {member.nama} berhasil diupdate!', 'success')
            return redirect(url_for('success', member_id=member.member_id))
        
        return render_template('edit.html', member=member)
    
    except Exception as e:
        print(f"Error in edit_member: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/perpanjangan', methods=['GET', 'POST'])
@login_required
def perpanjangan():
    member = None
    
    try:
        if request.method == 'POST':
            if 'search' in request.form:
                search_id = request.form.get('member_id', '').strip()
                member = Member.query.filter_by(member_id=search_id).first()
                
                if not member:
                    flash('ID Member tidak ditemukan!', 'error')
            
            elif 'extend' in request.form:
                member_id = request.form.get('member_id', '')
                member = Member.query.filter_by(member_id=member_id).first()
                
                if member:
                    tanggal_daftar_baru = member.tanggal_expire
                    tanggal_expire_baru = calculate_expire_date(tanggal_daftar_baru, member.type_member)
                    
                    member.tanggal_daftar = tanggal_daftar_baru
                    member.tanggal_expire = tanggal_expire_baru
                    
                    if member.type_member == 'bulanan':
                        member.total = member.biaya_bulanan
                    elif member.type_member == '3_bulan':
                        member.total = member.biaya_bulanan * 3
                    elif member.type_member == '6_bulan':
                        member.total = member.biaya_bulanan * 6
                    elif member.type_member == 'tahunan':
                        member.total = member.biaya_bulanan * 12
                    
                    db.session.commit()
                    
                    flash(f'Member {member.nama} berhasil diperpanjang!', 'success')
                    return redirect(url_for('success', member_id=member.member_id))
        
        return render_template('perpanjangan.html', member=member)
    
    except Exception as e:
        print(f"Error in perpanjangan: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/delete/<member_id>')
@login_required
def delete_confirm(member_id):
    try:
        member = Member.query.filter_by(member_id=member_id).first()
        
        if not member:
            flash('Member tidak ditemukan!', 'error')
            return redirect(url_for('index'))
        
        return render_template('delete.html', member=member)
    
    except Exception as e:
        print(f"Error in delete_confirm: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/delete/<member_id>/confirm', methods=['POST'])
@login_required
def delete_member(member_id):
    try:
        member = Member.query.filter_by(member_id=member_id).first()
        
        if not member:
            flash('Member tidak ditemukan!', 'error')
            return redirect(url_for('index'))
        
        nama = member.nama
        db.session.delete(member)
        db.session.commit()
        
        flash(f'Member {nama} (ID: {member_id}) berhasil dihapus!', 'success')
        return redirect(url_for('index'))
    
    except Exception as e:
        print(f"Error in delete_member: {e}")
        db.session.rollback()
        flash('Terjadi kesalahan saat menghapus member!', 'error')
        return redirect(url_for('index'))

@app.route('/profile')
@login_required
def profile():
    try:
        user = User.query.get(session['user_id'])
        return render_template('profile.html', user=user)
    except Exception as e:
        print(f"Error in profile: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/update-profile', methods=['POST'])
@login_required
def update_profile():
    try:
        user = User.query.get(session['user_id'])
        
        nama_lengkap = request.form.get('nama_lengkap')
        email = request.form.get('email')
        
        existing_email = User.query.filter(User.email == email, User.id != user.id).first()
        if existing_email:
            flash('Email sudah digunakan oleh user lain!', 'error')
            return redirect(url_for('profile'))
        
        user.nama_lengkap = nama_lengkap
        user.email = email
        
        session['nama_lengkap'] = nama_lengkap
        
        db.session.commit()
        
        flash('Profile berhasil diupdate!', 'success')
        return redirect(url_for('profile'))
    
    except Exception as e:
        print(f"Error in update_profile: {e}")
        flash('Terjadi kesalahan saat update profile!', 'error')
        return redirect(url_for('profile'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    try:
        if request.method == 'POST':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            user = User.query.get(session['user_id'])
            
            if not user.check_password(current_password):
                flash('Password lama tidak sesuai!', 'error')
                return render_template('change_password.html')
            
            if new_password != confirm_password:
                flash('Password baru tidak cocok!', 'error')
                return render_template('change_password.html')
            
            if len(new_password) < 6:
                flash('Password minimal 6 karakter!', 'error')
                return render_template('change_password.html')
            
            user.set_password(new_password)
            db.session.commit()
            
            flash('Password berhasil diubah!', 'success')
            return redirect(url_for('profile'))
        
        return render_template('change_password.html')
    
    except Exception as e:
        print(f"Error in change_password: {e}")
        flash('Terjadi kesalahan saat mengubah password!', 'error')
        return redirect(url_for('profile'))

@app.route('/success/<member_id>')
@login_required
def success(member_id):
    try:
        member = Member.query.filter_by(member_id=member_id).first()
        
        if not member:
            flash('Member tidak ditemukan!', 'error')
            return redirect(url_for('index'))
        
        qr_image = generate_qr_code_base64(member_id)
        auto_qr = request.args.get('auto_qr', '0')
        
        return render_template('success.html', 
                             member=member, 
                             qr_image=qr_image,
                             auto_qr=auto_qr)
    except Exception as e:
        print(f"Error in success: {e}")
        return f"<h1>Error: {str(e)}</h1><p><a href='/'>Back to Home</a></p>"

@app.route('/laporan')
@login_required
def laporan():
    try:
        all_members = Member.query.all()
        
        laporan_data = {}
        
        for member in all_members:
            bulan_tahun = member.tanggal_daftar.strftime('%Y-%m')
            bulan_label = member.tanggal_daftar.strftime('%B %Y')
            
            if bulan_tahun not in laporan_data:
                laporan_data[bulan_tahun] = {
                    'bulan_label': bulan_label,
                    'jumlah_member': 0,
                    'total_pendapatan': 0,
                    'member_list': []
                }
            
            laporan_data[bulan_tahun]['jumlah_member'] += 1
            laporan_data[bulan_tahun]['total_pendapatan'] += member.total
            laporan_data[bulan_tahun]['member_list'].append(member)
        
        laporan_sorted = dict(sorted(laporan_data.items(), reverse=True))
        
        total_keseluruhan = sum(m.total for m in all_members)
        jumlah_member_keseluruhan = len(all_members)
        
        return render_template('laporan.html', 
                             laporan_data=laporan_sorted,
                             total_keseluruhan=total_keseluruhan,
                             jumlah_member_keseluruhan=jumlah_member_keseluruhan)
    
    except Exception as e:
        print(f"Error in laporan: {e}")
        flash('Terjadi kesalahan saat memuat laporan!', 'error')
        return redirect(url_for('index'))

@app.route('/laporan/export')
@login_required
def export_laporan():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Pendapatan"
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="1F4E78")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "LAPORAN PENDAPATAN GYM"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = f"Tanggal: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='left')
        ws.merge_cells('A2:H2')
        
        current_row = 4
        
        all_members = Member.query.all()
        
        laporan_data = {}
        for member in all_members:
            bulan_tahun = member.tanggal_daftar.strftime('%Y-%m')
            bulan_label = member.tanggal_daftar.strftime('%B %Y')
            
            if bulan_tahun not in laporan_data:
                laporan_data[bulan_tahun] = {
                    'bulan_label': bulan_label,
                    'jumlah_member': 0,
                    'total_pendapatan': 0,
                    'member_list': []
                }
            
            laporan_data[bulan_tahun]['jumlah_member'] += 1
            laporan_data[bulan_tahun]['total_pendapatan'] += member.total
            laporan_data[bulan_tahun]['member_list'].append(member)
        
        laporan_sorted = dict(sorted(laporan_data.items(), reverse=True))
        
        grand_total = 0
        grand_member = 0
        
        for bulan_tahun, data in laporan_sorted.items():
            ws[f'A{current_row}'] = f"PERIODE: {data['bulan_label']}"
            ws[f'A{current_row}'].fill = section_fill
            ws[f'A{current_row}'].font = section_font
            ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 1
            
            headers = ['No', 'ID Member', 'Nama', 'Type Member', 'Tanggal Daftar', 'Biaya Bulanan', 'Biaya Pendaftaran', 'Total']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            current_row += 1
            
            for idx, member in enumerate(data['member_list'], 1):
                ws.cell(row=current_row, column=1).value = idx
                ws.cell(row=current_row, column=2).value = member.member_id
                ws.cell(row=current_row, column=3).value = member.nama
                ws.cell(row=current_row, column=4).value = member.type_member.replace('_', ' ').title()
                ws.cell(row=current_row, column=5).value = member.tanggal_daftar.strftime('%d-%m-%Y')
                ws.cell(row=current_row, column=6).value = member.biaya_bulanan
                ws.cell(row=current_row, column=7).value = member.biaya_pendaftaran
                ws.cell(row=current_row, column=8).value = member.total
                
                ws.cell(row=current_row, column=6).number_format = '#,##0'
                ws.cell(row=current_row, column=7).number_format = '#,##0'
                ws.cell(row=current_row, column=8).number_format = '#,##0'
                
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).border = border
                
                current_row += 1
            
            ws[f'A{current_row}'] = "TOTAL"
            ws[f'A{current_row}'].fill = total_fill
            ws[f'A{current_row}'].font = total_font
            ws[f'A{current_row}'].border = border
            
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
            
            ws[f'F{current_row}'].value = f"=SUM(F{current_row-len(data['member_list'])}:F{current_row-1})"
            ws[f'F{current_row}'].fill = total_fill
            ws[f'F{current_row}'].font = total_font
            ws[f'F{current_row}'].border = border
            ws[f'F{current_row}'].number_format = '#,##0'
            
            ws[f'G{current_row}'].value = f"=SUM(G{current_row-len(data['member_list'])}:G{current_row-1})"
            ws[f'G{current_row}'].fill = total_fill
            ws[f'G{current_row}'].font = total_font
            ws[f'G{current_row}'].border = border
            ws[f'G{current_row}'].number_format = '#,##0'
            
            ws[f'H{current_row}'].value = f"=SUM(H{current_row-len(data['member_list'])}:H{current_row-1})"
            ws[f'H{current_row}'].fill = total_fill
            ws[f'H{current_row}'].font = total_font
            ws[f'H{current_row}'].border = border
            ws[f'H{current_row}'].number_format = '#,##0'
            
            grand_total += data['total_pendapatan']
            grand_member += data['jumlah_member']
            
            current_row += 2
        
        current_row += 1
        ws[f'A{current_row}'] = "GRAND TOTAL KESELURUHAN"
        ws[f'A{current_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'].border = border
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        
        ws[f'H{current_row}'].value = grand_total
        ws[f'H{current_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws[f'H{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'H{current_row}'].border = border
        ws[f'H{current_row}'].number_format = '#,##0'
        
        current_row += 2
        ws[f'A{current_row}'] = "RINGKASAN"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        current_row += 1
        
        ws[f'A{current_row}'] = "Total Member:"
        ws[f'B{current_row}'] = grand_member
        current_row += 1
        
        ws[f'A{current_row}'] = "Total Pendapatan:"
        ws[f'B{current_row}'] = grand_total
        ws[f'B{current_row}'].number_format = '#,##0'
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Laporan_Pendapatan_{datetime.now().strftime('%d-%m-%Y_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error in export_laporan: {e}")
        flash('Terjadi kesalahan saat export laporan!', 'error')
        return redirect(url_for('laporan'))

@app.route('/qrcode/<member_id>')
@login_required
def generate_qrcode(member_id):
    try:
        member = Member.query.filter_by(member_id=member_id).first()
        
        if not member:
            flash('Member tidak ditemukan!', 'error')
            return redirect(url_for('index'))
        
        qr_image = generate_qr_code_base64(member_id)
        
        return render_template('qrcode.html', 
                             member=member, 
                             qr_image=qr_image)
    
    except Exception as e:
        print(f"Error in generate_qrcode: {e}")
        flash('Terjadi kesalahan saat generate QR Code!', 'error')
        return redirect(url_for('index'))

@app.route('/scan-public', methods=['GET', 'POST'])
def scan_public():
    try:
        if request.method == 'POST':
            qr_data = request.form.get('qr_data', '').strip()
            
            if qr_data.startswith('GYM-'):
                member_id = qr_data.replace('GYM-', '')
                member = Member.query.filter_by(member_id=member_id).first()
                
                if not member:
                    return render_template('scan_public.html', 
                                         error='Member tidak ditemukan!',
                                         qr_data=qr_data)
                
                today = datetime.now()
                if member.tanggal_expire < today:
                    return render_template('scan_public.html', 
                                         member=member, 
                                         expired=True,
                                         error=f'Member {member.nama} sudah expired!')
                
                last_attendance = Attendance.query.filter_by(
                    member_id=member_id
                ).order_by(Attendance.id.desc()).first()
                
                if last_attendance and last_attendance.status == 'check_in' and last_attendance.check_out is None:
                    last_attendance.check_out = datetime.now()
                    last_attendance.status = 'check_out'
                    db.session.commit()
                    
                    duration = last_attendance.check_out - last_attendance.check_in
                    hours = duration.total_seconds() / 3600
                    
                    return render_template('scan_public.html', 
                                         member=member, 
                                         action='check_out',
                                         attendance=last_attendance,
                                         duration=hours,
                                         success=f'Check-out berhasil! Terima kasih {member.nama}')
                else:
                    new_attendance = Attendance(
                        member_id=member_id,
                        check_in=datetime.now(),
                        status='check_in'
                    )
                    db.session.add(new_attendance)
                    db.session.commit()
                    
                    return render_template('scan_public.html', 
                                         member=member, 
                                         action='check_in',
                                         attendance=new_attendance,
                                         success=f'Check-in berhasil! Selamat datang {member.nama}')
            else:
                return render_template('scan_public.html', 
                                     error='QR Code tidak valid! Format harus GYM-XXXXXX')
        
        return render_template('scan_public.html')
    
    except Exception as e:
        print(f"Error in scan_public: {e}")
        return render_template('scan_public.html', 
                             error='Terjadi kesalahan sistem!')

@app.route('/scan', methods=['GET', 'POST'])
@login_required
def scan_qrcode():
    try:
        if request.method == 'POST':
            qr_data = request.form.get('qr_data', '').strip()
            
            if qr_data.startswith('GYM-'):
                member_id = qr_data.replace('GYM-', '')
                member = Member.query.filter_by(member_id=member_id).first()
                
                if not member:
                    flash('Member tidak ditemukan!', 'error')
                    return redirect(url_for('scan_qrcode'))
                
                today = datetime.now()
                if member.tanggal_expire < today:
                    flash(f'Member {member.nama} sudah expired! Silakan perpanjang membership.', 'error')
                    return render_template('scan.html', member=member, expired=True)
                
                last_attendance = Attendance.query.filter_by(
                    member_id=member_id
                ).order_by(Attendance.id.desc()).first()
                
                if last_attendance and last_attendance.status == 'check_in' and last_attendance.check_out is None:
                    last_attendance.check_out = datetime.now()
                    last_attendance.status = 'check_out'
                    db.session.commit()
                    
                    flash(f'Check-out berhasil! Terima kasih {member.nama}', 'success')
                    return render_template('scan.html', 
                                         member=member, 
                                         action='check_out',
                                         attendance=last_attendance)
                else:
                    new_attendance = Attendance(
                        member_id=member_id,
                        check_in=datetime.now(),
                        status='check_in'
                    )
                    db.session.add(new_attendance)
                    db.session.commit()
                    
                    flash(f'Check-in berhasil! Selamat datang {member.nama}', 'success')
                    return render_template('scan.html', 
                                         member=member, 
                                         action='check_in',
                                         attendance=new_attendance)
            else:
                flash('QR Code tidak valid!', 'error')
        
        return render_template('scan.html')
    
    except Exception as e:
        print(f"Error in scan_qrcode: {e}")
        flash('Terjadi kesalahan saat scan QR Code!', 'error')
        return redirect(url_for('scan_qrcode'))

@app.route('/attendance')
@login_required
def attendance_list():
    try:
        filter_date = request.args.get('date', '')
        filter_member = request.args.get('member', '')
        
        query = Attendance.query
        
        if filter_date:
            try:
                target_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
                query = query.filter(db.func.date(Attendance.check_in) == target_date)
            except:
                pass
        
        if filter_member:
            query = query.filter(Attendance.member_id.like(f'%{filter_member}%'))
        
        attendances = query.order_by(Attendance.check_in.desc()).all()
        
        today = datetime.now().date()
        today_attendances = Attendance.query.filter(
            db.func.date(Attendance.check_in) == today
        ).all()
        
        checked_in_today = len([a for a in today_attendances if a.check_out is None])
        total_today = len(today_attendances)
        
        return render_template('attendance.html', 
                             attendances=attendances,
                             checked_in_today=checked_in_today,
                             total_today=total_today,
                             filter_date=filter_date,
                             filter_member=filter_member)
    
    except Exception as e:
        print(f"Error in attendance_list: {e}")
        flash('Terjadi kesalahan saat memuat data attendance!', 'error')
        return redirect(url_for('index'))

@app.route('/attendance/export')
@login_required
def export_attendance():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Kehadiran"
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "LAPORAN KEHADIRAN MEMBER GYM"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = f"Tanggal Export: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='left')
        
        current_row = 4
        headers = ['No', 'ID Member', 'Nama', 'Check-in', 'Check-out', 'Durasi (jam)', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        current_row += 1
        
        attendances = Attendance.query.order_by(Attendance.check_in.desc()).all()
        
        for idx, attendance in enumerate(attendances, 1):
            ws.cell(row=current_row, column=1).value = idx
            ws.cell(row=current_row, column=2).value = attendance.member_id
            ws.cell(row=current_row, column=3).value = attendance.member.nama
            ws.cell(row=current_row, column=4).value = attendance.check_in.strftime('%d-%m-%Y %H:%M:%S')
            
            if attendance.check_out:
                ws.cell(row=current_row, column=5).value = attendance.check_out.strftime('%d-%m-%Y %H:%M:%S')
                duration = (attendance.check_out - attendance.check_in).total_seconds() / 3600
                ws.cell(row=current_row, column=6).value = round(duration, 2)
                ws.cell(row=current_row, column=7).value = "Check-out"
            else:
                ws.cell(row=current_row, column=5).value = "-"
                ws.cell(row=current_row, column=6).value = "-"
                ws.cell(row=current_row, column=7).value = "Check-in"
            
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Data_Kehadiran_{datetime.now().strftime('%d-%m-%Y_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error in export_attendance: {e}")
        flash('Terjadi kesalahan saat export data!', 'error')
        return redirect(url_for('attendance_list'))

@app.route('/telegram/check-expiry')
@login_required
def telegram_check_expiry():
    """Manual trigger untuk check member yang akan expired"""
    try:
        check_expiring_members(app)
        flash('Telegram alert berhasil dikirim!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/telegram/test')
@login_required
def telegram_test():
    """Test Telegram bot connection"""
    try:
        bot = get_telegram_bot()
        if bot:
            today = datetime.now()
            member_aktif = Member.query.filter(Member.tanggal_expire >= today).count()
            member_expired = Member.query.filter(Member.tanggal_expire < today).count()
            total_pendapatan = db.session.query(db.func.sum(Member.total)).scalar() or 0
            
            tiga_hari_kedepan = today + timedelta(days=3)
            akan_expired = Member.query.filter(
                Member.tanggal_expire >= today,
                Member.tanggal_expire <= tiga_hari_kedepan
            ).count()
            
            stats = {
                'member_aktif': member_aktif,
                'member_expired': member_expired,
                'akan_expired': akan_expired,
                'total_pendapatan': total_pendapatan
            }
            
            bot.send_daily_summary(stats)
            flash('Test message berhasil dikirim ke Telegram!', 'success')
        else:
            flash('Telegram bot belum dikonfigurasi!', 'error')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('index'))


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    # 🤖 Start Telegram Bot in background thread
    import threading
    from telegram_bot import run_telegram_bot
    
    def start_bot():
        try:
            run_telegram_bot(app)
        except Exception as e:
            print(f"❌ Bot error: {e}")
    
    bot_thread = threading.Thread(target=start_bot, daemon=True)
    bot_thread.start()
    
    print("⏳ Waiting 2 seconds for bot to initialize...")
    import time
    time.sleep(2)
    
    # Start Flask app
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)



